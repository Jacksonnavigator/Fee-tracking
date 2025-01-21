import smtplib
from email.mime.text import MIMEText
from models import Student, Session
from datetime import datetime
from openpyxl import load_workbook
import streamlit as st

def update_fee_status(session, student_id, amount_paid):
    student = session.query(Student).get(student_id)
    if student:
        try:
            student.fee_due -= amount_paid
            student.last_payment = datetime.now()
            student.is_paid = student.fee_due <= 0
            session.commit()
            return True
        except Exception as e:
            session.rollback()
            st.error(f"An error occurred while updating fee: {str(e)}")
    return False

def get_student(session, student_id):
    return session.query(Student).get(student_id)

def send_notification(student_email, message):
    msg = MIMEText(message)
    msg['Subject'] = "Fee Notification"
    msg['From'] = "your-email@example.com"
    msg['To'] = student_email

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login("your-email@example.com", "your-password")  # Use env variables in production
            server.send_message(msg)
        print("Successfully sent email")
    except Exception as e:
        print(f"Failed to send email: {e}")

def import_students_from_excel(session, file_path):
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    students_added = []
    students_updated = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming header in row 1
        if all(row):  # Check if any cell in the row is not empty
            try:
                name, email, fee_due = row[:3]  # Assuming these are the first three columns
                existing_student = session.query(Student).filter_by(email=email).first()
                
                if existing_student:
                    existing_student.name = name  # Update name if you want, or leave as is
                    existing_student.fee_due = float(fee_due)  # Convert to float to ensure numeric value
                    students_updated.append(email)
                else:
                    new_student = Student(name=name, email=email, fee_due=float(fee_due))
                    session.add(new_student)
                    students_added.append(email)
            except ValueError:
                st.error(f"Error processing row for {email}. Check data format.")
            except Exception as e:
                st.error(f"Unexpected error for {email}: {str(e)}")
    
    try:
        session.commit()
        st.success(f"Import successful. Added {len(students_added)} new student(s), Updated {len(students_updated)} student(s).")
        return True
    except Exception as e:
        session.rollback()
        st.error(f"Failed to commit changes to database: {str(e)}")
        return False
